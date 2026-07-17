"""
Module J: PBM_Markup_Analysis.xlsx workbook builder.

Generates the human-browsable 4-sheet Excel workbook (PBM Markup, States,
Methodology, Sources) from `output/leaderboard.csv` and
`output/state_summary.csv` -- those two CSVs are the canonical
machine-readable record; this workbook is a rendering of them, not a
separate analysis. Previously hand-maintained with no committed build
script (see `dataset/REPRODUCE.md`, section 6); this replaces that ad hoc
step with a deterministic one.

Named `i_workbook` was requested but `modules/i_unpriced_drugs.py`
already claims the `i_` slot in this suite's alphabetical module
sequence, so this uses the next free letter, `j_`.
"""
from __future__ import annotations

import datetime
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import config  # noqa: E402

LEADERBOARD_PATH = config.OUTPUT_DIR / "leaderboard.csv"
STATE_SUMMARY_PATH = config.OUTPUT_DIR / "state_summary.csv"

_NAVY = "0A0E1A"
_DARK_GREEN = "006400"
_PURPLE = "660099"
_DARK_RED = "8B0000"
_ORANGE = "E67300"
_RED = "CC0000"
_LIGHT_GREY = "F2F2F2"
_LIGHT_YELLOW = "FFF9C4"
_BLANK_PLACEHOLDER = "—"  # em dash, for a citation-derived cell with no value

_TITLE_FONT = Font(bold=True, size=16, color=_NAVY)
_SECTION_FONT = Font(bold=True, size=14, color=_NAVY)
_DATA_HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
_DATA_HEADER_FILL = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_SOURCES_HEADER_FONT = Font(bold=True, size=11, color=_NAVY)
_SOURCES_HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_ROW_SHADE_FILL = PatternFill(start_color=_LIGHT_GREY, end_color=_LIGHT_GREY, fill_type="solid")
_HIGHLIGHT_FILL = PatternFill(start_color=_LIGHT_YELLOW, end_color=_LIGHT_YELLOW, fill_type="solid")
_MARKUP_FONT = Font(bold=True, color=_ORANGE)
_PRICE_FONT = Font(bold=True, color=_RED)
_TOP10_BORDER = Border(bottom=Side(style="thick"))
_MARKUP_NUMBER_FORMAT = '#,##0.00"%"'
_CURRENCY_FORMAT = '"$"#,##0.00'

# Source types have different selection properties (see METHODOLOGY.md,
# "Public citation enrichment") -- color-coding them on the sheet itself
# means a reader isn't left to assume a federal industry study and a
# plaintiff's litigation exhibit carry the same evidentiary weight.
_SOURCE_TYPE_COLORS = {
    "federal_study": _NAVY,
    "state_disclosure": _DARK_GREEN,
    "peer_reviewed": _PURPLE,
    "litigation": _DARK_RED,
}

_SOURCE_TYPE_LEGEND = [
    ("Source types (see METHODOLOGY.md for the selection-bias caveat):", None),
    ("federal_study — FTC interim reports (industry-wide sample)", "federal_study"),
    ("state_disclosure — Maine MHDO (mandated reporting)", "state_disclosure"),
    ("peer_reviewed — JAMA Mattingly (academic sample)", "peer_reviewed"),
    ("litigation — J&J and Wells Fargo ERISA complaints (plaintiff-selected, not representative)", "litigation"),
]

_PBM_MARKUP_COLUMNS = [
    ("rank", "Rank"),
    ("drug_term", "Drug"),
    ("costplus_per_unit", "Cost Plus/unit"),
    ("nadac_per_unit", "NADAC Acquisition/unit"),
    ("partd_per_unit", "Medicare Pays/unit"),
    ("gap_partd", "Gap/unit"),
    ("total_overpayment", "Total Overpayment"),
    ("best_confirmed_spread", "Confirmed Markup %"),
    ("estimated_pbm_price_per_unit", "Estimated PBM Price/unit"),
    ("best_confirmed_source", "Source Citation"),
    ("source_type", "Source Type"),
    ("canonical_unit", "Unit"),
]
_PBM_MARKUP_HEADER_ROW = 10
_TOP_N_BOLD_ROWS = 10
_MIN_COL_WIDTH = 8
_MAX_COL_WIDTH = 70

_STATE_COLUMNS = [
    ("state", "State"),
    ("total_medicaid_overpayment", "Total Medicaid Overpayment"),
    ("top_drug", "Top Drug"),
    ("top_drug_overpayment", "Top Drug Overpayment"),
    ("drugs_analyzed", "Drugs Analyzed"),
]
_STATE_HEADER_ROW = 3

_SOURCES_HEADERS = ["Source Name", "Year", "Publisher", "URL", "What it contains"]
_SOURCES_COL_WIDTHS = [43, 12, 32, 22, 60]
_SOURCES_ROWS = [
    ("FTC Second Interim Staff Report on PBMs", "2025", "FTC", "ftc.gov",
     "51 specialty generic drugs with confirmed markup percentages"),
    ("FTC First Interim Staff Report on PBMs", "2024", "FTC", "ftc.gov",
     "Two cancer drugs, $1.6B excess revenue"),
    ("Maine MHDO Drug Price Transparency Report", "2024", "Maine Health Data Organization", "mhdo.maine.gov",
     "Drug-level WAC and PBM reimbursement"),
    ("Ohio Auditor PBM Spread Report", "2018", "Ohio Auditor of State", "ohioauditor.gov",
     "$224.8M spread on Medicaid generics"),
    ("JAMA Health Forum Mattingly et al.", "2023", "JAMA", "jamanetwork.com",
     "45 high-utilization generics with PBM gross profit breakdown"),
    ("CMS NADAC", "Weekly", "CMS", "data.medicaid.gov", "Drug acquisition costs"),
    ("CMS Medicare Part D Spending by Drug", "Annual", "CMS", "data.cms.gov", "Medicare per-drug spending"),
    ("CMS Medicaid SDUD", "Quarterly", "CMS", "data.medicaid.gov",
     "State-level Medicaid drug utilization and reimbursement"),
    ("Lewandowski v. Johnson & Johnson, ERISA Complaint", "2024", "D.N.J., Civil No. 1:24-cv-00671",
     "courtlistener.com/docket/68223269",
     "42-drug table: NADAC acquisition cost vs. price J&J's plans paid Express Scripts"),
    ("Navarro v. Wells Fargo & Co., ERISA Complaint", "2024", "D. Minn., Civil No. 0:24-cv-03043",
     "courtlistener.com/docket/68995654",
     "38-drug table: NADAC acquisition cost vs. price Wells Fargo's plan paid Express Scripts"),
    ('46brooklyn Research, "Wrecklimid"', "2026", "46brooklyn Research", "46brooklyn.com/research",
     "Abiraterone: ESI-affiliate vs. non-affiliate pharmacy premium, Georgia commercial NADAC disclosure filings"),
]

# (line, is_section_header) -- section headers get _SECTION_FONT; everything
# else is a plain row (including the None spacer rows between sections).
# The drug count in the first bullet is filled in at build time from the
# actual leaderboard row count -- see _methodology_lines(). The specific
# "88.1% of the Cost Plus catalog matched" figure that used to appear here
# and in the last bullet is not derivable from leaderboard.csv/
# state_summary.csv alone, so it is described qualitatively instead of
# re-asserting a number this script cannot verify.


def _methodology_lines(n_drugs: int) -> list[tuple[str | None, bool]]:
    return [
        ("What this analysis covers", True),
        (f"This workbook quantifies the estimated annual overpayment by Medicare Part D and Medicaid on generic "
         f"drugs, compared to what the same drugs cost at Cost Plus Drugs. It covers {n_drugs:,} matched generic "
         f"drugs.", False),
        (None, False),
        ("Data sources", True),
        ("•  Cost Plus Drugs prices: Cost Plus GraphQL storefront API", False),
        ("•  NADAC (National Average Drug Acquisition Cost): CMS, updated weekly. The actual price pharmacies "
         "pay to acquire each drug.", False),
        ("•  Medicare Part D Spending by Drug: CMS annual dataset.", False),
        ("•  Medicaid State Drug Utilization Data (SDUD): CMS quarterly dataset, broken out by state.", False),
        ("•  Confirmed spread citations: federal reports including FTC Second Interim Staff Report (Jan 2025), "
         "Maine MHDO Drug Price Transparency Report, Ohio Auditor PBM Spread Report (2018), JAMA Health Forum "
         "(Mattingly et al., Oct 2023), and others. See Source Citation column.", False),
        (None, False),
        ("Methodology", True),
        ("•  All comparisons are on a per-unit basis (per tablet, per mL, per gram) using the NADAC Pricing "
         "Unit as the canonical unit.", False),
        ("•  Cost Plus per-unit price = (acquisition cost × 1.15 + pharmacy fee) ÷ package "
         "quantity.", False),
        ("•  Overpayment = (Medicare or Medicaid price per unit − Cost Plus price per unit) × total "
         "dosage units. As of this version, Medicare Part D's overpayment is dollarized ONCE PER MOLECULE (the "
         "highest-priced strength represents the molecule; see “Part D molecule-level fix” below), not "
         "once per strength.", False),
        ("•  Only generic drugs are included. Brand drug rebates are negotiated privately and never "
         "disclosed; including them would require estimating numbers that do not exist publicly.", False),
        ("•  Net prices are never estimated. The \"net_per_unit\" field is always \"not public\" by "
         "design.", False),
        ("•  Estimated PBM Price: where a confirmed markup percentage exists from a named government or "
         "academic source, an estimated PBM billing price is computed as NADAC acquisition cost × (1 + markup "
         "%). This is an estimate derived from a confirmed markup percentage, not a directly observed transaction "
         "price. It is labeled explicitly as estimated throughout.", False),
        (None, False),
        ("Part D molecule-level fix", True),
        ("•  Medicare Part D publishes spending by generic name, not by strength -- one national "
         "Tot_Dsg_Unts/Tot_Spndng figure covers every strength of a molecule combined. Multiplying that same "
         "national figure by each strength's own per-unit gap and summing across every strength row -- the prior "
         "behavior -- counted the national total once per strength (e.g. Atorvastatin's 4 strengths each carried "
         "the full national unit count).", False),
        ("•  Fixed by dollarizing once per molecule: the strength with the HIGHEST Cost Plus per-unit price "
         "represents the molecule (minimizes the gap, so the figure is a conservative floor, never inflated). "
         "Every other strength row of that molecule contributes $0. Per-strength Gap/unit stays real and visible "
         "for every row regardless.", False),
        (None, False),
        ("Honest limitations", True),
        ("•  Medicare Part D spending figures are gross of manufacturer rebates. For generics, rebates are "
         "minimal; for brands, this analysis excludes them entirely.", False),
        ("•  Medicaid SDUD reimbursement rates are a proxy for commercial reimbursement, not an exact "
         "equivalent.", False),
        ("•  Shipping fees ($5 per order at Cost Plus) are excluded from per-unit calculations as they are "
         "per-order, not per-unit.", False),
        ("•  Match confidence: matched to federal drug codes via the NIH RxNav crosswalk; unmatched drugs are "
         "excluded from this workbook.", False),
        ("•  No public data gives the real strength-mix Part D dispenses at, so the highest-price-strength "
         "rule is a deliberate, documented choice, not an estimate of the true weighted-average price.", False),
    ]


def _snapshot_date_label(leaderboard_path: Path) -> str:
    """Human-readable date this workbook's source data was built, taken
    from leaderboard.csv's own mtime rather than hardcoded, so it never
    goes stale the way the hand-maintained workbook's header text did."""
    dt = datetime.datetime.fromtimestamp(leaderboard_path.stat().st_mtime)
    return f"{dt:%B} {dt.day}, {dt:%Y}"


def _set_column_widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


_CURRENCY_COLS = {
    "costplus_per_unit", "nadac_per_unit", "partd_per_unit", "gap_partd",
    "total_overpayment", "estimated_pbm_price_per_unit",
}


def _display_str(value, col_name: str) -> str:
    """Estimated rendered width of a cell, used only for autofit sizing --
    matches the number_format actually applied (2-decimal currency /
    percent), not Python's raw float repr, so autofit doesn't over-widen
    columns based on floating-point noise nobody sees on screen."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if col_name in _CURRENCY_COLS and isinstance(value, (int, float)):
        return f"${value:,.2f}"
    if col_name == "best_confirmed_spread" and isinstance(value, (int, float)):
        return f"{value:,.2f}%"
    return str(value)


def _autofit_column_widths(ws, col_count: int, header_row: int, last_row: int) -> None:
    for c in range(1, col_count + 1):
        max_len = 0
        for r in range(header_row, last_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, max_len + 2))


def _build_pbm_markup_sheet(wb: Workbook, leaderboard: pd.DataFrame, snapshot_date: str) -> None:
    ws = wb.active
    ws.title = "PBM Markup"
    ws["A1"] = "PBM Markup Analysis — Generic Drug Overpayment vs Cost Plus Prices"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "Estimated annual Medicare Part D + Medicaid overpayment. Generics only. Net prices never estimated."
    ws["A3"] = f"Data as of {snapshot_date} (leaderboard.csv build time)"
    for offset, (line, source_type) in enumerate(_SOURCE_TYPE_LEGEND):
        cell = ws.cell(row=4 + offset, column=1, value=line)
        if source_type is not None:
            cell.font = Font(bold=True, color=_SOURCE_TYPE_COLORS[source_type])

    for col_idx, (_, header) in enumerate(_PBM_MARKUP_COLUMNS, start=1):
        cell = ws.cell(row=_PBM_MARKUP_HEADER_ROW, column=col_idx, value=header)
        cell.font = _DATA_HEADER_FONT
        cell.fill = _DATA_HEADER_FILL

    markup_col = [i for i, (name, _) in enumerate(_PBM_MARKUP_COLUMNS, start=1) if name == "best_confirmed_spread"][0]
    price_col = [i for i, (name, _) in enumerate(_PBM_MARKUP_COLUMNS, start=1)
                 if name == "estimated_pbm_price_per_unit"][0]
    source_type_col = [i for i, (name, _) in enumerate(_PBM_MARKUP_COLUMNS, start=1) if name == "source_type"][0]
    total_overpayment_col = [i for i, (name, _) in enumerate(_PBM_MARKUP_COLUMNS, start=1)
                              if name == "total_overpayment"][0]

    # Only these two are citation-derived (present for 56 of 2,062 rows) --
    # "—" marks "no confirmed public citation for this drug", distinct from
    # Excel's plain empty cell used elsewhere for values that are simply
    # not applicable (e.g. a molecule-level Part D row -- see "Part D
    # molecule-level fix" in METHODOLOGY.md).
    _DASH_PLACEHOLDER_COLS = {"best_confirmed_spread", "estimated_pbm_price_per_unit"}

    last_row = _PBM_MARKUP_HEADER_ROW + len(leaderboard)
    for data_idx, (_, row) in enumerate(leaderboard.iterrows(), start=1):
        r = _PBM_MARKUP_HEADER_ROW + data_idx
        is_top_n = data_idx <= _TOP_N_BOLD_ROWS
        for c, (col_name, _) in enumerate(_PBM_MARKUP_COLUMNS, start=1):
            value = row[col_name]
            blank = pd.isna(value)
            if blank:
                cell_value = _BLANK_PLACEHOLDER if col_name in _DASH_PLACEHOLDER_COLS else None
            else:
                cell_value = value
            cell = ws.cell(row=r, column=c, value=cell_value)

            font_color = None
            font_bold = is_top_n
            if col_name == "best_confirmed_spread":
                cell.number_format = _MARKUP_NUMBER_FORMAT
                if not blank:
                    font_color, font_bold = _ORANGE, True
            elif col_name == "estimated_pbm_price_per_unit":
                cell.number_format = _CURRENCY_FORMAT
                if not blank:
                    font_color, font_bold = _RED, True
            elif col_name in _CURRENCY_COLS:
                cell.number_format = _CURRENCY_FORMAT
            elif col_name == "source_type" and not blank:
                font_color = _SOURCE_TYPE_COLORS.get(value)

            if font_color or font_bold:
                cell.font = Font(bold=font_bold, color=font_color)

            # Alternate row shading (every other DATA row, independent of
            # the top-10 bolding/border below) makes a 2,000+ row sheet
            # readable without a filter applied.
            if data_idx % 2 == 0:
                cell.fill = _ROW_SHADE_FILL

            if data_idx == _TOP_N_BOLD_ROWS:
                cell.border = _TOP10_BORDER

    ws.conditional_formatting.add(
        f"{get_column_letter(total_overpayment_col)}{_PBM_MARKUP_HEADER_ROW + 1}:"
        f"{get_column_letter(total_overpayment_col)}{last_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color=_DARK_RED),
    )

    _autofit_column_widths(ws, len(_PBM_MARKUP_COLUMNS), _PBM_MARKUP_HEADER_ROW, last_row)
    ws.freeze_panes = f"A{_PBM_MARKUP_HEADER_ROW + 1}"


_STATE_HIGHLIGHT = {"CA", "NY", "OH"}
_STATE_CURRENCY_COLS = {"total_medicaid_overpayment", "top_drug_overpayment"}


def _build_states_sheet(wb: Workbook, state_summary: pd.DataFrame) -> None:
    ws = wb.create_sheet("States")
    ws["A1"] = "State-Level Medicaid Overpayment vs Cost Plus Prices"
    ws["A1"].font = _TITLE_FONT

    for col_idx, (_, header) in enumerate(_STATE_COLUMNS, start=1):
        cell = ws.cell(row=_STATE_HEADER_ROW, column=col_idx, value=header)
        cell.font = _DATA_HEADER_FONT
        cell.fill = _DATA_HEADER_FILL

    last_row = _STATE_HEADER_ROW + len(state_summary)
    for r, (_, row) in enumerate(state_summary.iterrows(), start=_STATE_HEADER_ROW + 1):
        highlight = row["state"] in _STATE_HIGHLIGHT
        for c, (col_name, _) in enumerate(_STATE_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=row[col_name])
            if col_name in _STATE_CURRENCY_COLS:
                cell.number_format = _CURRENCY_FORMAT
            if highlight:
                cell.fill = _HIGHLIGHT_FILL

    _autofit_column_widths(ws, len(_STATE_COLUMNS), _STATE_HEADER_ROW, last_row)
    ws.freeze_panes = "A4"


def _build_methodology_sheet(wb: Workbook, n_drugs: int) -> None:
    ws = wb.create_sheet("Methodology")
    for r, (line, is_header) in enumerate(_methodology_lines(n_drugs), start=1):
        cell = ws.cell(row=r, column=1, value=line)
        if is_header:
            cell.font = _SECTION_FONT
    ws.column_dimensions["A"].width = 110


def _build_sources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Sources")
    for c, header in enumerate(_SOURCES_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = _SOURCES_HEADER_FONT
        cell.fill = _SOURCES_HEADER_FILL
    for r, row in enumerate(_SOURCES_ROWS, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    _set_column_widths(ws, _SOURCES_COL_WIDTHS)


def run(
    leaderboard_path: Path | None = None,
    state_summary_path: Path | None = None,
) -> Workbook:
    """Builds and returns the 4-sheet workbook. Does not write to disk --
    see report.write_workbook() for that, matching the CSV modules'
    build/write split."""
    leaderboard_path = leaderboard_path or LEADERBOARD_PATH
    state_summary_path = state_summary_path or STATE_SUMMARY_PATH
    leaderboard = pd.read_csv(leaderboard_path)
    state_summary = pd.read_csv(state_summary_path)
    snapshot_date = _snapshot_date_label(leaderboard_path)

    wb = Workbook()
    wb.properties.title = "PBM Markup Analysis"
    wb.properties.creator = "Agastya Dhar"
    _build_pbm_markup_sheet(wb, leaderboard, snapshot_date)
    _build_states_sheet(wb, state_summary)
    _build_methodology_sheet(wb, len(leaderboard))
    _build_sources_sheet(wb)

    print(f"[module_j] built PBM_Markup_Analysis.xlsx: {len(leaderboard):,} drug rows, "
          f"{len(state_summary):,} state rows")
    return wb


if __name__ == "__main__":
    result = run()
    out_path = config.OUTPUT_DIR / "PBM_Markup_Analysis.xlsx"
    result.save(out_path)
    print(f"[module_j] wrote -> {out_path}")
